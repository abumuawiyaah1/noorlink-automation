#!/usr/bin/env python3
"""
Generate NoorLink breakage-fulfillment country routing from WeConnect P1 per-MB pricelist.

Usage:
  python scripts/generate_breakage_strategy.py
  python scripts/generate_breakage_strategy.py --xlsx "/path/to/WEC eSIM P1 Global per MB....xlsx"
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from dataclasses import asdict, dataclass
from datetime import date
from pathlib import Path
from statistics import median
from typing import Any, Dict, List, Optional, Tuple

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

DEFAULT_XLSX = Path.home() / "Downloads" / "WEC eSIM P1 Global per MB 17 JUNE 26 1.1 2.xlsx"
OUT_DIR = ROOT / "data" / "breakage"

WECONNECT_ESIM_USD = 1.60
MIN_MARGIN_AT_FULL_USAGE = 0.25  # require 25% margin if customer uses 100% of 10GB plan

# Reference retail anchors (competitive bundle sticker prices)
RETAIL_10GB_USD = 29.99
RETAIL_3GB_USD = 19.99
RETAIL_1GB_USD = 14.99

# Standard analysis plan for margin gate
GATE_PLAN_GB = 10
GATE_PLAN_MB = GATE_PLAN_GB * 1024

USAGE_SCENARIOS = (0.25, 0.50, 0.75, 1.0)

SAUDI_NAMES = {"saudi arabia"}

CARIBBEAN_ISLANDS = {
    "anguilla",
    "antigua and barbuda",
    "aruba",
    "bahamas",
    "barbados",
    "bermuda",
    "bonaire",
    "british virgin islands",
    "cayman islands",
    "cuba",
    "curacao",
    "dominica",
    "dominican republic",
    "grenada",
    "guadeloupe",
    "haiti",
    "jamaica",
    "martinique",
    "montserrat",
    "puerto rico",
    "saint kitts and nevis",
    "saint lucia",
    "saint vincent and the grenadines",
    "sint maarten",
    "trinidad and tobago",
    "turks and caicos islands",
    "us virgin islands",
}

LATAM_TELNA_COUNTRIES = {
    "argentina",
    "bolivia",
    "brazil",
    "chile",
    "colombia",
    "costa rica",
    "ecuador",
    "el salvador",
    "guatemala",
    "honduras",
    "mexico",
    "nicaragua",
    "panama",
    "paraguay",
    "peru",
    "uruguay",
    "venezuela",
}

EXCLUDE_NAMES = {
    "satellite networks",
}


def slugify(value: str) -> str:
    lowered = value.strip().lower()
    lowered = re.sub(r"[^a-z0-9]+", "-", lowered)
    return lowered.strip("-")


def norm_country(value: str) -> str:
    return re.sub(r"\s+", " ", value.strip().lower())


@dataclass
class CountryRow:
    country: str
    country_slug: str
    operator: str
    price_mb_usd: float
    price_gb_usd: float
    policy: str
    policy_reason: str
    margin_10gb_100pct: float
    margin_10gb_50pct: float
    margin_3gb_100pct: float
    breakage_score: int
    region_hint: str


def load_weconnect_rows(xlsx: Path) -> List[Dict[str, Any]]:
    try:
        import openpyxl
    except ImportError:
        raise SystemExit("Install openpyxl: pip install openpyxl")

    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    rows: List[Dict[str, Any]] = []
    for r in range(8, ws.max_row + 1):
        country = ws.cell(r, 1).value
        operator = ws.cell(r, 2).value
        price = ws.cell(r, 8).value
        if not country or price is None:
            continue
        try:
            price_mb = float(price)
        except (TypeError, ValueError):
            continue
        rows.append(
            {
                "country": str(country).strip(),
                "operator": str(operator or "").strip(),
                "price_mb_usd": price_mb,
            }
        )
    return rows


def cheapest_by_country(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    out: Dict[str, Dict[str, Any]] = {}
    for row in rows:
        key = norm_country(row["country"])
        prev = out.get(key)
        if prev is None or row["price_mb_usd"] < prev["price_mb_usd"]:
            out[key] = row
    return out


def wholesale_data_usd(price_mb: float, data_gb: float, usage_pct: float) -> float:
    mb = data_gb * 1024 * usage_pct
    return price_mb * mb


def margin_usd(retail: float, wholesale: float) -> float:
    return retail - wholesale


def classify_country(
    country_name: str,
    price_mb: float,
    operator: str,
) -> CountryRow:
    key = norm_country(country_name)
    slug = slugify(country_name)

    price_gb = price_mb * 1024
    w10_100 = wholesale_data_usd(price_mb, GATE_PLAN_GB, 1.0) + WECONNECT_ESIM_USD
    w10_50 = wholesale_data_usd(price_mb, GATE_PLAN_GB, 0.5) + WECONNECT_ESIM_USD
    w3_100 = wholesale_data_usd(price_mb, 3, 1.0) + WECONNECT_ESIM_USD

    m10_100 = margin_usd(RETAIL_10GB_USD, w10_100)
    m10_50 = margin_usd(RETAIL_10GB_USD, w10_50)
    m3_100 = margin_usd(RETAIL_3GB_USD, w3_100)

    region_hint = "global"
    if key in SAUDI_NAMES:
        region_hint = "middle-east"
    elif key in CARIBBEAN_ISLANDS:
        region_hint = "caribbean"
    elif key in LATAM_TELNA_COUNTRIES:
        region_hint = "south-america"
    elif price_gb < 1.5:
        region_hint = "europe-tier"
    elif price_gb < 3.0:
        region_hint = "standard"

    policy = "catalog_cascade"
    reason = "Default: use existing provider catalog / Access / Telna cascade."

    if key in EXCLUDE_NAMES:
        policy = "exclude"
        reason = "Non-terrestrial or non-viable destination."
    elif key in SAUDI_NAMES:
        policy = "access_fixed"
        reason = "Saudi/Umrah policy: eSIM Access fixed bundles only."
    elif key in CARIBBEAN_ISLANDS:
        policy = "telna_fixed"
        reason = "Caribbean per-MB wholesale too high; Telna regional bundles required."
    elif m10_100 < 0:
        policy = "telna_fixed" if key in LATAM_TELNA_COUNTRIES else "catalog_cascade"
        reason = (
            "100% usage on 10GB plan loses money on WeConnect per-MB; "
            "route to fixed bundle provider."
        )
    elif m10_100 >= RETAIL_10GB_USD * MIN_MARGIN_AT_FULL_USAGE:
        policy = "weconnect_breakage"
        reason = (
            f"10GB full-usage margin ${m10_100:.2f} (≥25% target). "
            "Eligible for virtual bundle on WeConnect PAYG."
        )
    elif m3_100 >= 0 and key in LATAM_TELNA_COUNTRIES:
        policy = "telna_fixed"
        reason = "LatAm: prefer Telna fixed bundles until usage telemetry confirms breakage."
    elif m3_100 >= RETAIL_3GB_USD * 0.15:
        policy = "weconnect_breakage"
        reason = (
            f"10GB gate tight but 3GB full-usage margin ${m3_100:.2f}; "
            "breakage on smaller plans only."
        )
    elif price_gb >= 5.0:
        policy = "catalog_cascade"
        reason = f"High per-GB (${price_gb:.2f}); do not use breakage until rates improve."

    # Breakage score 0-100 for sorting pilot countries
    if policy == "weconnect_breakage":
        score = int(min(100, max(0, (m10_50 / RETAIL_10GB_USD) * 100)))
    elif policy == "access_fixed":
        score = 0
    elif policy == "telna_fixed":
        score = 5
    elif policy == "exclude":
        score = 0
    else:
        score = int(min(40, max(0, (m10_100 / RETAIL_10GB_USD) * 100)))

    return CountryRow(
        country=country_name,
        country_slug=slug,
        operator=operator,
        price_mb_usd=round(price_mb, 8),
        price_gb_usd=round(price_gb, 4),
        policy=policy,
        policy_reason=reason,
        margin_10gb_100pct=round(m10_100, 2),
        margin_10gb_50pct=round(m10_50, 2),
        margin_3gb_100pct=round(m3_100, 2),
        breakage_score=score,
        region_hint=region_hint,
    )


def build_summary(countries: List[CountryRow]) -> Dict[str, Any]:
    by_policy: Dict[str, int] = {}
    for row in countries:
        by_policy[row.policy] = by_policy.get(row.policy, 0) + 1

    pilot = sorted(
        [c for c in countries if c.policy == "weconnect_breakage"],
        key=lambda r: (-r.breakage_score, r.price_gb_usd),
    )[:25]

    med_gb = median([c.price_gb_usd for c in countries])
    return {
        "generated_on": date.today().isoformat(),
        "source": "WeConnect P1 Global per MB",
        "weconnect_esim_usd": WECONNECT_ESIM_USD,
        "retail_anchors_usd": {
            "1gb": RETAIL_1GB_USD,
            "3gb": RETAIL_3GB_USD,
            "10gb": RETAIL_10GB_USD,
        },
        "gate_plan_gb": GATE_PLAN_GB,
        "min_margin_at_full_usage_pct": MIN_MARGIN_AT_FULL_USAGE,
        "country_count": len(countries),
        "median_price_gb_usd": round(med_gb, 4),
        "policy_counts": by_policy,
        "pilot_countries_top_25": [asdict(c) for c in pilot],
    }


def write_csv(path: Path, countries: List[CountryRow]) -> None:
    fields = list(asdict(countries[0]).keys()) if countries else []
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=fields)
        writer.writeheader()
        for row in countries:
            writer.writerow(asdict(row))


def write_margin_scenarios(path: Path, countries: List[CountryRow]) -> None:
    plans = [
        ("1gb_5d", 1, RETAIL_1GB_USD),
        ("3gb_7d", 3, RETAIL_3GB_USD),
        ("10gb_15d", 10, RETAIL_10GB_USD),
    ]
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(
            [
                "country_slug",
                "country",
                "policy",
                "price_gb_usd",
                "plan",
                "retail_usd",
                "usage_pct",
                "wholesale_usd",
                "margin_usd",
            ]
        )
        for row in countries:
            for plan_key, gb, retail in plans:
                for usage in USAGE_SCENARIOS:
                    wholesale = (
                        wholesale_data_usd(row.price_mb_usd, gb, usage)
                        + WECONNECT_ESIM_USD
                    )
                    writer.writerow(
                        [
                            row.country_slug,
                            row.country,
                            row.policy,
                            row.price_gb_usd,
                            plan_key,
                            retail,
                            usage,
                            round(wholesale, 2),
                            round(retail - wholesale, 2),
                        ]
                    )


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate breakage strategy data files.")
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX)
    parser.add_argument("--out", type=Path, default=OUT_DIR)
    args = parser.parse_args()

    if not args.xlsx.is_file():
        print(f"Missing pricelist: {args.xlsx}", file=sys.stderr)
        return 1

    raw = load_weconnect_rows(args.xlsx)
    cheapest = cheapest_by_country(raw)
    countries = sorted(
        [
            classify_country(row["country"], row["price_mb_usd"], row["operator"])
            for row in cheapest.values()
        ],
        key=lambda r: r.country,
    )

    args.out.mkdir(parents=True, exist_ok=True)
    routing = {
        "version": 1,
        "summary": build_summary(countries),
        "countries": {row.country_slug: asdict(row) for row in countries},
    }
    (args.out / "country_routing.json").write_text(
        json.dumps(routing, indent=2), encoding="utf-8"
    )
    write_csv(args.out / "country_routing.csv", countries)
    write_margin_scenarios(args.out / "margin_scenarios.csv", countries)

    print(f"Wrote {len(countries)} countries to {args.out}")
    for policy, count in sorted(routing["summary"]["policy_counts"].items()):
        print(f"  {policy}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
